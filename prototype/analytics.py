"""K12Online report parsing and anonymous aggregation.

This module deliberately knows nothing about FastAPI, SQLAlchemy or provider SDKs.  It accepts
CSV/XLSX export rows, rejects unsupported/unsafe input, and yields only normalized fields.
"""
from __future__ import annotations

import csv
import hashlib
import hmac
import io
import math
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook

MAX_ANALYTICS_REPORT_BYTES = 10 * 1024 * 1024
MAX_ANALYTICS_ROWS = 50_000
SUPPORTED_REPORT_TYPES = {"text/csv", "application/vnd.ms-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}

HEADER_ALIASES = {
    "learner_identifier": {"learner_pseudonym", "student_id", "student_code", "ma_hoc_sinh", "ma_hoc_vien"},
    "course_external_id": {"course_external_id", "course_id", "ma_khoa_hoc", "ma_chuong_trinh"},
    "course_title": {"course_title", "course_name", "ten_khoa_hoc", "ten_chuong_trinh"},
    "class_code": {"class_code", "class_id", "ma_lop"},
    "lesson_external_id": {"lesson_external_id", "lesson_id", "activity_id", "ma_bai_hoc", "ma_hoat_dong"},
    "lesson_title": {"lesson_title", "lesson_name", "activity_title", "ten_bai_hoc", "ten_hoat_dong"},
    "activity_date": {"activity_date", "activity_at", "last_activity_at", "ngay_hoat_dong"},
    "duration_minutes": {"duration_minutes", "duration", "thoi_luong_phut"},
    "completion_percent": {"completion_percent", "completion_ratio", "progress", "ti_le_hoan_thanh"},
    "completion_status": {"completion_status", "status", "trang_thai_hoan_thanh"},
    "score": {"score", "diem"},
    "max_score": {"max_score", "diem_toi_da"},
    "attempt_number": {"attempt_number", "attempt", "lan_lam"},
    "correct_answers": {"correct_answers", "so_cau_dung"},
    "total_questions": {"total_questions", "question_count", "tong_so_cau"},
    "correct_rate": {"correct_rate", "correct_ratio", "ti_le_dung"},
}
REQUIRED_FIELDS = {"learner_identifier", "course_external_id", "lesson_external_id"}


class AnalyticsImportError(ValueError):
    pass


def _header(value: Any) -> str:
    return "_".join(str(value or "").strip().casefold().replace("đ", "d").replace(" ", "_").split("_"))


def discover_mapping(headers: list[Any]) -> dict[str, str]:
    available = {_header(header): str(header).strip() for header in headers if str(header or "").strip()}
    mapping: dict[str, str] = {}
    for field, aliases in HEADER_ALIASES.items():
        matching = next((alias for alias in aliases if alias in available), None)
        if matching:
            mapping[field] = available[matching]
    missing = sorted(REQUIRED_FIELDS - set(mapping))
    if missing:
        raise AnalyticsImportError(
            "Thiếu cột bắt buộc: " + ", ".join(missing) + ". Hãy dùng mã học viên giả danh, mã khóa học và mã bài học; không dùng họ tên/email học sinh."
        )
    return mapping


def parse_report(filename: str, content_type: str | None, raw: bytes) -> tuple[list[dict[str, Any]], dict[str, str]]:
    if not raw:
        raise AnalyticsImportError("Tệp báo cáo trống.")
    if len(raw) > MAX_ANALYTICS_REPORT_BYTES:
        raise AnalyticsImportError("Tệp báo cáo vượt quá 10 MB.")
    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if suffix == "csv":
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise AnalyticsImportError("CSV phải dùng mã hóa UTF-8.") from exc
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        mapping = discover_mapping(headers)
        rows = list(reader)
    elif suffix == "xlsx":
        if content_type and content_type not in SUPPORTED_REPORT_TYPES:
            raise AnalyticsImportError("Loại tệp XLSX không hợp lệ.")
        try:
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = list(next(values, ()))
            mapping = discover_mapping(headers)
            header_names = [str(value).strip() if value is not None else "" for value in headers]
            rows = [dict(zip(header_names, row)) for row in values if any(value is not None and str(value).strip() for value in row)]
            workbook.close()
        except AnalyticsImportError:
            raise
        except Exception as exc:
            raise AnalyticsImportError("Không thể đọc XLSX. Hãy dùng tệp báo cáo .xlsx hợp lệ.") from exc
    else:
        raise AnalyticsImportError("Chỉ nhận báo cáo CSV UTF-8 hoặc XLSX.")
    if not rows:
        raise AnalyticsImportError("Báo cáo không có dòng dữ liệu.")
    if len(rows) > MAX_ANALYTICS_ROWS:
        raise AnalyticsImportError("Báo cáo vượt quá 50.000 dòng; hãy chia nhỏ báo cáo.")
    return rows, mapping


def _clean_text(value: Any, *, max_length: int = 300) -> str | None:
    if value is None:
        return None
    cleaned = " ".join(str(value).split())
    if not cleaned:
        return None
    if len(cleaned) > max_length:
        raise AnalyticsImportError("Một giá trị văn bản vượt quá giới hạn cho phép.")
    return cleaned


def _number(value: Any, *, field: str, whole: bool = False, minimum: float | None = 0) -> float | int | None:
    if value in (None, ""):
        return None
    try:
        parsed = float(str(value).replace(",", ".").strip())
    except (TypeError, ValueError) as exc:
        raise AnalyticsImportError(f"{field} phải là số.") from exc
    if not math.isfinite(parsed) or (minimum is not None and parsed < minimum):
        raise AnalyticsImportError(f"{field} không hợp lệ.")
    if whole:
        if not parsed.is_integer():
            raise AnalyticsImportError(f"{field} phải là số nguyên.")
        return int(parsed)
    return parsed


def _ratio(value: Any, *, field: str) -> float | None:
    parsed = _number(value, field=field, minimum=0)
    if parsed is None:
        return None
    ratio = float(parsed)
    if ratio > 1:
        ratio /= 100
    if ratio > 1:
        raise AnalyticsImportError(f"{field} phải nằm trong khoảng 0–100%.")
    return ratio


def _datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=value.tzinfo or timezone.utc)
    text = str(value).strip().replace("Z", "+00:00")
    for candidate in (text, text.replace("/", "-")):
        try:
            parsed = datetime.fromisoformat(candidate)
            return parsed.replace(tzinfo=parsed.tzinfo or timezone.utc)
        except ValueError:
            pass
    for layout in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, layout).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    raise AnalyticsImportError("activity_date không đúng định dạng ngày/giờ.")


def normalize_rows(rows: list[dict[str, Any]], mapping: dict[str, str], *, pseudonym_key: bytes) -> tuple[list[dict[str, Any]], dict[str, int]]:
    accepted: list[dict[str, Any]] = []
    errors: dict[str, int] = {}
    for row_number, row in enumerate(rows, start=2):
        try:
            def item(field: str) -> Any:
                header = mapping.get(field)
                return row.get(header) if header else None

            learner_identifier = _clean_text(item("learner_identifier"), max_length=200)
            course_external_id = _clean_text(item("course_external_id"), max_length=200)
            lesson_external_id = _clean_text(item("lesson_external_id"), max_length=200)
            if not learner_identifier or not course_external_id or not lesson_external_id:
                raise AnalyticsImportError("Thiếu định danh giả danh, mã khóa học hoặc mã bài học.")
            score = _number(item("score"), field="score")
            max_score = _number(item("max_score"), field="max_score")
            if score is not None and max_score is not None and score > max_score:
                raise AnalyticsImportError("score không được lớn hơn max_score.")
            correct_answers = _number(item("correct_answers"), field="correct_answers", whole=True)
            total_questions = _number(item("total_questions"), field="total_questions", whole=True)
            if correct_answers is not None and total_questions is not None and correct_answers > total_questions:
                raise AnalyticsImportError("correct_answers không được lớn hơn total_questions.")
            correct_ratio = _ratio(item("correct_rate"), field="correct_rate")
            if correct_ratio is None and correct_answers is not None and total_questions:
                correct_ratio = int(correct_answers) / int(total_questions)
            token = hmac.new(pseudonym_key, learner_identifier.encode("utf-8"), hashlib.sha256).hexdigest()
            accepted.append({
                "row_number": row_number,
                "learner_token": token,
                "course_external_id": course_external_id,
                "course_title": _clean_text(item("course_title")),
                "class_code": _clean_text(item("class_code"), max_length=100),
                "lesson_external_id": lesson_external_id,
                "lesson_title": _clean_text(item("lesson_title")),
                "activity_at": _datetime(item("activity_date")),
                "duration_minutes": _number(item("duration_minutes"), field="duration_minutes"),
                "completion_ratio": _ratio(item("completion_percent"), field="completion_percent"),
                "completion_status": _clean_text(item("completion_status"), max_length=40),
                "score": score,
                "max_score": max_score,
                "attempt_number": _number(item("attempt_number"), field="attempt_number", whole=True, minimum=1),
                "correct_answers": correct_answers,
                "total_questions": total_questions,
                "correct_ratio": correct_ratio,
            })
        except AnalyticsImportError as exc:
            key = str(exc)
            errors[key] = errors.get(key, 0) + 1
    if not accepted:
        raise AnalyticsImportError("Không có dòng nào hợp lệ để nhập. Dữ liệu gốc không được lưu.")
    return accepted, errors


def aggregate_events(events: list[Any]) -> dict[str, Any]:
    """Return school-wide aggregates only; no learner-level result leaves the service."""
    def average(values: list[float]) -> float | None:
        return round(sum(values) / len(values), 4) if values else None

    completion = [float(event.completion_ratio) for event in events if event.completion_ratio is not None]
    duration = [float(event.duration_minutes) for event in events if event.duration_minutes is not None]
    score_ratios = [float(event.score) / float(event.max_score) for event in events if event.score is not None and event.max_score not in (None, 0)]
    correct = [float(event.correct_ratio) for event in events if event.correct_ratio is not None]
    groups: dict[tuple[str, str | None], list[Any]] = {}
    for event in events:
        groups.setdefault((event.lesson_external_id, event.lesson_title), []).append(event)
    lessons = []
    for (lesson_id, lesson_title), group in sorted(groups.items(), key=lambda pair: pair[0][0]):
        lessons.append({
            "lesson_external_id": lesson_id,
            "lesson_title": lesson_title,
            "event_count": len(group),
            "completion_ratio": average([float(event.completion_ratio) for event in group if event.completion_ratio is not None]),
            "score_ratio": average([float(event.score) / float(event.max_score) for event in group if event.score is not None and event.max_score not in (None, 0)]),
            "correct_ratio": average([float(event.correct_ratio) for event in group if event.correct_ratio is not None]),
        })
    return {
        "event_count": len(events),
        "learner_count": len({event.learner_token for event in events}),
        "completion_ratio": average(completion),
        "score_ratio": average(score_ratios),
        "correct_ratio": average(correct),
        "average_duration_minutes": average(duration),
        "lessons": lessons,
    }


def aggregate_insights(summary: dict[str, Any]) -> list[str]:
    """Safe initial substitute for an LLM: deterministic guidance from anonymous aggregates."""
    insights: list[str] = []
    if summary["event_count"] == 0:
        return ["Chưa có dữ liệu đã nhập. Hãy nhập báo cáo K12Online đã ẩn danh."]
    if summary["completion_ratio"] is not None and summary["completion_ratio"] < 0.8:
        insights.append("Tỷ lệ hoàn thành dưới 80%; nên kiểm tra độ dài bài, điểm dừng và hướng dẫn học trước khi thay đổi nội dung.")
    if summary["correct_ratio"] is not None and summary["correct_ratio"] < 0.65:
        insights.append("Tỷ lệ đúng dưới 65%; hãy xem lại các bài có chỉ số thấp và bổ sung ví dụ hoặc câu hỏi luyện tập.")
    if summary["score_ratio"] is not None and summary["score_ratio"] < 0.65:
        insights.append("Điểm trung bình dưới 65%; giáo viên nên rà soát mục tiêu, phản hồi sau quiz và mức độ câu hỏi.")
    if not insights:
        insights.append("Các chỉ số tổng hợp hiện chưa cho thấy cảnh báo rõ rệt; giáo viên vẫn cần đối chiếu với bối cảnh lớp học.")
    return insights
